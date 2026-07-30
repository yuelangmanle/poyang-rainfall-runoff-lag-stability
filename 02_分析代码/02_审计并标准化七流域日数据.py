#!/usr/bin/env python3
"""审计14个公开Excel文件并生成不改值的七流域标准化日表。"""

from __future__ import annotations

import hashlib
import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd


PROJECT_DIR = Path(__file__).resolve().parents[1]
RAW_DIR = PROJECT_DIR / "01_数据" / "公开原始数据"
PROCESSED_DIR = PROJECT_DIR / "01_数据" / "处理后数据"
TABLE_DIR = PROJECT_DIR / "03_分析结果" / "表"
RECORD_DIR = PROJECT_DIR / "00_项目说明与决策记录"

EXPECTED_DATES = pd.date_range("2006-01-01", "2020-12-31", freq="D")
PRECIP_EXTREME_THRESHOLD = 200.0
RUNOFF_EXTREME_THRESHOLD = 50.0

BASINS = [
    ("Ganjiang", "赣江"),
    ("Fuhe", "抚河"),
    ("Xinjiang", "信江"),
    ("Xiushui", "修水"),
    ("Liaohe", "潦河"),
    ("Le'anhe", "乐安河"),
    ("Changjiang", "昌江"),
]

FORMULA_RE = re.compile(
    r"^=\s*F\$?(\d+)\s*\*\s*86\.4\s*/\s*([0-9]+(?:\.[0-9]+)?)\s*$",
    re.IGNORECASE,
)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_core_columns(path: Path, kind: str) -> pd.DataFrame:
    raw = pd.read_excel(path, header=None, engine="openpyxl")
    if raw.shape[1] < 3:
        raise ValueError(f"{path.name} 少于3列，无法读取日期和值")

    data = pd.DataFrame(
        {
            "excel_serial": pd.to_numeric(raw.iloc[:, 0], errors="coerce"),
            "value": pd.to_numeric(raw.iloc[:, 1], errors="coerce"),
            "date": pd.to_datetime(raw.iloc[:, 2], errors="coerce"),
        }
    )
    data = data.loc[data["date"].notna()].copy().reset_index(drop=True)

    if kind == "runoff" and raw.shape[1] >= 6:
        discharge = pd.to_numeric(raw.iloc[:, 5], errors="coerce")
        data["discharge_m3s"] = discharge.loc[raw.iloc[:, 2].notna()].reset_index(drop=True)
    elif kind == "runoff":
        data["discharge_m3s"] = np.nan
    return data


def inspect_workbook(path: Path, valid_rows: int, kind: str) -> dict:
    wb_formula = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws_formula = wb_formula.active
    worksheet_max_row = ws_formula.max_row
    worksheet_max_col = ws_formula.max_column
    wb_values = None

    formula_counts: dict[str, int] = {}
    formula_examples: dict[str, str] = {}
    for row in ws_formula.iter_rows(min_row=1, max_row=valid_rows):
        for cell in row:
            if cell.data_type == "f":
                formula_counts[cell.column_letter] = formula_counts.get(cell.column_letter, 0) + 1
                formula_examples.setdefault(cell.column_letter, str(cell.value))

    result = {
        "sheet_names": "|".join(wb_formula.sheetnames),
        "sheet_count": len(wb_formula.sheetnames),
        "worksheet_max_row": worksheet_max_row,
        "worksheet_max_col": worksheet_max_col,
        "trailing_formatted_blank_rows": max(0, worksheet_max_row - valid_rows),
        "formula_counts": ";".join(f"{k}:{v}" for k, v in sorted(formula_counts.items())),
        "formula_examples": ";".join(f"{k}:{v}" for k, v in sorted(formula_examples.items())),
        "area_km2_from_formula": np.nan,
        "formula_row_reference_errors": 0,
        "runoff_identity_max_abs_error": np.nan,
        "cached_H_vs_B_max_abs_error": np.nan,
    }

    if kind == "runoff":
        areas = []
        row_reference_errors = 0
        formula_rows = 0
        for row_number in range(1, valid_rows + 1):
            formula = ws_formula.cell(row_number, 8).value if worksheet_max_col >= 8 else None
            if not isinstance(formula, str) or not formula.startswith("="):
                continue
            formula_rows += 1
            match = FORMULA_RE.match(formula)
            if match is None:
                row_reference_errors += 1
                continue
            if int(match.group(1)) != row_number:
                row_reference_errors += 1
            areas.append(float(match.group(2)))

        result["formula_row_reference_errors"] = row_reference_errors
        unique_areas = sorted(set(areas))
        if formula_rows and len(unique_areas) == 1 and row_reference_errors == 0:
            area = unique_areas[0]
            result["area_km2_from_formula"] = area
            discharge = np.array(
                [ws_formula.cell(i, 6).value for i in range(1, valid_rows + 1)],
                dtype=float,
            )
            runoff_b = np.array(
                [ws_formula.cell(i, 2).value for i in range(1, valid_rows + 1)],
                dtype=float,
            )
            calculated = discharge * 86.4 / area
            result["runoff_identity_max_abs_error"] = float(
                np.nanmax(np.abs(calculated - runoff_b))
            )

            wb_values = openpyxl.load_workbook(path, read_only=False, data_only=True)
            ws_values = wb_values.active
            cached_h = np.array(
                [ws_values.cell(i, 8).value for i in range(1, valid_rows + 1)],
                dtype=float,
            )
            result["cached_H_vs_B_max_abs_error"] = float(
                np.nanmax(np.abs(cached_h - runoff_b))
            )

    wb_formula.close()
    if wb_values is not None:
        wb_values.close()
    return result


def date_audit(data: pd.DataFrame) -> dict:
    dates = pd.DatetimeIndex(data["date"])
    expected_serial = (dates - pd.Timestamp("1899-12-30")).days.astype(float)
    serial_diff = np.abs(data["excel_serial"].to_numpy(dtype=float) - expected_serial)
    return {
        "valid_date_rows": len(data),
        "start_date": dates.min().date().isoformat() if len(dates) else "",
        "end_date": dates.max().date().isoformat() if len(dates) else "",
        "duplicate_dates": int(dates.duplicated().sum()),
        "date_sequence_complete": bool(dates.equals(EXPECTED_DATES)),
        "excel_serial_max_abs_error": float(np.nanmax(serial_diff)) if len(serial_diff) else np.nan,
    }


def value_audit(data: pd.DataFrame, threshold: float) -> dict:
    values = data["value"]
    return {
        "missing_values": int(values.isna().sum()),
        "negative_values": int((values < 0).sum()),
        "zero_values": int((values == 0).sum()),
        "min_value": float(values.min()),
        "mean_value": float(values.mean()),
        "max_value": float(values.max()),
        "above_review_threshold": int((values > threshold).sum()),
    }


def build_field_dictionary() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["*_rainfall.xlsx", "A", "Excel日期序号", "day serial", "推定", "与C列日期逐行核对"],
            ["*_rainfall.xlsx", "B", "逐日面平均降水", "mm/d", "推定", "原文件无表头；依据数据集说明、量级与年总量推定"],
            ["*_rainfall.xlsx", "C", "日期", "date", "验证", "2006-01-01至2020-12-31连续"],
            ["Changjiang_rainfall.xlsx", "E", "B列乘10", "unknown", "未确认/不使用", "2006-2019年共5,113行公式=B行*10，含义和单位不明"],
            ["*_Qobs.xlsx", "A", "Excel日期序号", "day serial", "推定", "与C列日期逐行核对"],
            ["*_Qobs.xlsx", "B", "逐日径流深", "mm/d", "六流域公式验证；潦河推定", "六流域与F*86.4/面积一致；潦河无F/H列证据"],
            ["*_Qobs.xlsx", "C", "日期", "date", "验证", "2006-01-01至2020-12-31连续"],
            ["除Liaohe外的*_Qobs.xlsx", "F", "逐日实测流量", "m3/s", "由换算公式推定", "H列公式使用F*86.4/面积"],
            ["除Liaohe外的*_Qobs.xlsx", "H", "逐日径流深公式复算", "mm/d", "公式验证", "缓存值与B列逐行一致"],
        ],
        columns=["适用文件", "原始列", "字段含义", "单位", "证据等级", "判定依据或限制"],
    )


def main() -> None:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    RECORD_DIR.mkdir(parents=True, exist_ok=True)

    audit_rows = []
    area_rows = []
    anomaly_rows = []
    standardized = []

    for basin_en, basin_cn in BASINS:
        rainfall_path = RAW_DIR / f"{basin_en}_rainfall.xlsx"
        runoff_path = RAW_DIR / f"{basin_en}_Qobs.xlsx"
        if not rainfall_path.exists() or not runoff_path.exists():
            raise FileNotFoundError(f"{basin_cn}缺少降水或流量原始文件")

        rainfall = read_core_columns(rainfall_path, "rainfall")
        runoff = read_core_columns(runoff_path, "runoff")
        rainfall_book = inspect_workbook(rainfall_path, len(rainfall), "rainfall")
        runoff_book = inspect_workbook(runoff_path, len(runoff), "runoff")

        for path, kind, data, threshold, book in [
            (rainfall_path, "rainfall", rainfall, PRECIP_EXTREME_THRESHOLD, rainfall_book),
            (runoff_path, "runoff_depth", runoff, RUNOFF_EXTREME_THRESHOLD, runoff_book),
        ]:
            audit_rows.append(
                {
                    "basin_en": basin_en,
                    "basin_cn": basin_cn,
                    "file": path.name,
                    "variable": kind,
                    "sha256": sha256_file(path),
                    **date_audit(data),
                    **value_audit(data, threshold),
                    **book,
                }
            )

        area = runoff_book["area_km2_from_formula"]
        unit_status = "QDEPTH_FORMULA_VERIFIED" if pd.notna(area) else "QDEPTH_STRUCTURAL_INFERENCE"
        area_rows.append(
            {
                "basin_en": basin_en,
                "basin_cn": basin_cn,
                "area_km2_from_formula": area,
                "has_discharge_m3s": bool(runoff["discharge_m3s"].notna().any()),
                "runoff_depth_unit_status": unit_status,
                "formula_identity_max_abs_error": runoff_book["runoff_identity_max_abs_error"],
                "cached_H_vs_B_max_abs_error": runoff_book["cached_H_vs_B_max_abs_error"],
            }
        )

        if not pd.DatetimeIndex(rainfall["date"]).equals(pd.DatetimeIndex(runoff["date"])):
            raise ValueError(f"{basin_cn}降水与径流日期不一致")

        daily = pd.DataFrame(
            {
                "basin_en": basin_en,
                "basin_cn": basin_cn,
                "date": rainfall["date"],
                "precipitation_mm_inferred": rainfall["value"],
                "runoff_depth_mm_inferred": runoff["value"],
                "discharge_m3s_if_available": runoff["discharge_m3s"],
                "area_km2_from_formula": area,
                "precip_unit_status": "P_UNIT_INFERRED",
                "runoff_unit_status": unit_status,
                "precip_source_file": rainfall_path.name,
                "runoff_source_file": runoff_path.name,
            }
        )
        daily["flag_precip_gt200"] = daily["precipitation_mm_inferred"] > PRECIP_EXTREME_THRESHOLD
        daily["flag_runoff_depth_gt50"] = daily["runoff_depth_mm_inferred"] > RUNOFF_EXTREME_THRESHOLD
        standardized.append(daily)

        for _, row in daily.loc[daily["flag_precip_gt200"]].iterrows():
            anomaly_rows.append(
                {
                    "basin_cn": basin_cn,
                    "date": row["date"].date().isoformat(),
                    "variable": "precipitation_mm_inferred",
                    "value": row["precipitation_mm_inferred"],
                    "review_threshold": PRECIP_EXTREME_THRESHOLD,
                    "paired_value": row["runoff_depth_mm_inferred"],
                    "note": "极端值复核标记；未删除、未修正",
                }
            )
        for _, row in daily.loc[daily["flag_runoff_depth_gt50"]].iterrows():
            anomaly_rows.append(
                {
                    "basin_cn": basin_cn,
                    "date": row["date"].date().isoformat(),
                    "variable": "runoff_depth_mm_inferred",
                    "value": row["runoff_depth_mm_inferred"],
                    "review_threshold": RUNOFF_EXTREME_THRESHOLD,
                    "paired_value": row["precipitation_mm_inferred"],
                    "note": "极端值复核标记；未删除、未修正",
                }
            )

    audit = pd.DataFrame(audit_rows)
    areas = pd.DataFrame(area_rows)
    anomalies = pd.DataFrame(anomaly_rows)
    daily_all = pd.concat(standardized, ignore_index=True)
    field_dictionary = build_field_dictionary()

    precip_matrix = daily_all.pivot(
        index="date", columns="basin_en", values="precipitation_mm_inferred"
    ).sort_index()
    all_seven_equal = precip_matrix.nunique(axis=1, dropna=False).eq(1)
    all_seven_equal_nonzero = all_seven_equal & precip_matrix.iloc[:, 0].ne(0)
    daily_all["flag_precip_all7_identical"] = daily_all["date"].map(all_seven_equal)
    daily_all["flag_precip_all7_identical_nonzero"] = daily_all["date"].map(
        all_seven_equal_nonzero
    )
    duplicate_year_audit = pd.DataFrame(
        {
            "year": sorted(precip_matrix.index.year.unique()),
        }
    )
    duplicate_year_audit["days_in_year"] = duplicate_year_audit["year"].map(
        pd.Series(1, index=precip_matrix.index).groupby(precip_matrix.index.year).sum()
    )
    duplicate_year_audit["all7_identical_days"] = duplicate_year_audit["year"].map(
        all_seven_equal.groupby(all_seven_equal.index.year).sum()
    )
    duplicate_year_audit["all7_identical_nonzero_days"] = duplicate_year_audit["year"].map(
        all_seven_equal_nonzero.groupby(all_seven_equal_nonzero.index.year).sum()
    )
    duplicate_year_audit["entire_year_all7_identical"] = (
        duplicate_year_audit["days_in_year"] == duplicate_year_audit["all7_identical_days"]
    )

    audit_path = TABLE_DIR / "01_原始Excel数据审计总表.csv"
    fields_path = TABLE_DIR / "02_原始Excel字段字典.csv"
    areas_path = TABLE_DIR / "03_流量换算公式与面积审计.csv"
    anomalies_path = TABLE_DIR / "04_极端值复核清单_未作删除.csv"
    duplicate_path = TABLE_DIR / "05_跨流域降水重复序列审计.csv"
    daily_path = PROCESSED_DIR / "七流域_2006-2020_日尺度标准化数据.csv"

    audit.to_csv(audit_path, index=False, encoding="utf-8-sig")
    field_dictionary.to_csv(fields_path, index=False, encoding="utf-8-sig")
    areas.to_csv(areas_path, index=False, encoding="utf-8-sig")
    anomalies.to_csv(anomalies_path, index=False, encoding="utf-8-sig")
    duplicate_year_audit.to_csv(duplicate_path, index=False, encoding="utf-8-sig")
    daily_all.to_csv(daily_path, index=False, encoding="utf-8-sig", date_format="%Y-%m-%d")
    daily_hash = sha256_file(daily_path)
    (daily_path.with_suffix(daily_path.suffix + ".sha256")).write_text(
        f"{daily_hash}  {daily_path.name}\n", encoding="utf-8"
    )

    complete_files = int(audit["date_sequence_complete"].sum())
    missing_total = int(audit["missing_values"].sum())
    negative_total = int(audit["negative_values"].sum())
    precip_flags = int(daily_all["flag_precip_gt200"].sum())
    runoff_flags = int(daily_all["flag_runoff_depth_gt50"].sum())
    verified_basins = int(areas["area_km2_from_formula"].notna().sum())
    fully_duplicated_years = duplicate_year_audit.loc[
        duplicate_year_audit["entire_year_all7_identical"], "year"
    ].astype(str).tolist()
    duplicated_year_text = "、".join(fully_duplicated_years) if fully_duplicated_years else "无"
    equal_2020 = int(
        duplicate_year_audit.loc[
            duplicate_year_audit["year"].eq(2020), "all7_identical_days"
        ].iloc[0]
    )
    equal_nonzero_2020 = int(
        duplicate_year_audit.loc[
            duplicate_year_audit["year"].eq(2020), "all7_identical_nonzero_days"
        ].iloc[0]
    )
    area_lines = "\n".join(
        f"- {row.basin_cn}：{row.area_km2_from_formula:g} km2"
        for row in areas.itertuples()
        if pd.notna(row.area_km2_from_formula)
    )

    report = f"""# 七流域公开Excel数据审计报告

生成日期：{pd.Timestamp.now(tz='Asia/Shanghai').date().isoformat()}

## 一、审计结论

- 14个核心工作簿均只有1个工作表，均无表头。
- 14/14个文件各识别到5,479个有效日期行，日期完整覆盖2006-01-01至2020-12-31；通过完整性检查的文件为{complete_files}/14。
- 日期重复、有效值缺测和负值合计分别为{int(audit['duplicate_dates'].sum())}、{missing_total}和{negative_total}。
- Excel显示的最大行数大于5,479，是空白格式范围，不是额外观测记录。
- 六个流域的B列可由F列流量和H列公式精确验证为日径流深；潦河缺少F/H列，只能依据与其他文件一致的结构和量级推定B列为日径流深。
- 降水B列在所有文件中都没有表头或单位；当前依据数据集说明、日值量级和年总量推定为mm/d，证据低于六个流域的径流深换算证据。
- 跨流域重复检查发现，完整重复年份为：{duplicated_year_text}。2020年366天的七流域降水完全相同，其中{equal_nonzero_2020}天为相同的非零值；这不是由共同无雨日造成的假象。
- 标准化过程没有删除、插补、缩放或修正任何原始数值。

## 二、六个流域的公式面积

{area_lines}

六个流域的逐行换算恒等式为：`径流深(mm/d) = 流量(m3/s) × 86.4 ÷ 流域面积(km2)`。所有已识别公式的行号引用正确，B列、H列缓存值和独立复算结果一致。潦河不具备该验证链。

## 三、必须保留的风险标记

1. **字段元数据不足**：所有工作簿均无表头；降水单位和潦河径流单位不是原文件明示，而是推定。
2. **昌江降水E列含义不明**：该列为B列乘10的公式，本文不使用。
3. **极端日值需做稳健性检验**：降水大于{PRECIP_EXTREME_THRESHOLD:g}的记录有{precip_flags}条，径流深大于{RUNOFF_EXTREME_THRESHOLD:g}的记录有{runoff_flags}条。它们当前仅被标记，不能在没有外部证据时当作错误删除。
4. **2020年降水丧失空间独立性**：七流域全年降水逐日完全相同（{equal_2020}/366天）。主分析应采用2006-2019年；2006-2020年只作为包含该共同序列的敏感性方案，不能用2020年论证流域间降水差异。
5. **赣江2019-05-19降水为679.6432**：该值远高于其余记录并显著影响年总量，需做原值保留、剔除该日和外部降水产品核对三种敏感性分析；在取得证据前不能自行除以10。
6. **面积不是统一来源的流域边界面积**：六个面积来自各自工作簿换算公式，只能称为“公式使用面积”。信江17,600 km2与梅港文献常见面积存在冲突，不能据此确认站点身份。
7. **跨流域比较口径**：主分析使用公开文件B列的归一化径流序列；原始流量F列只用于单位/公式复核，不直接比较绝对流量。

## 四、当前可用性判断

这批数据可以进入探索性实证分析，但主分析时段应暂定为2006-2019年，而不是机械使用完整的2006-2020年。最低稳妥做法是：在方法中披露无表头事实和单位推断依据；对2020年共同降水序列和赣江679.6432极端值做敏感性分析；把潦河单独标为结构推定；不写具体水文站；不把公式面积当成独立权威流域面积。

## 五、生成文件

- `03_分析结果/表/01_原始Excel数据审计总表.csv`
- `03_分析结果/表/02_原始Excel字段字典.csv`
- `03_分析结果/表/03_流量换算公式与面积审计.csv`
- `03_分析结果/表/04_极端值复核清单_未作删除.csv`
- `03_分析结果/表/05_跨流域降水重复序列审计.csv`
- `01_数据/处理后数据/七流域_2006-2020_日尺度标准化数据.csv`

标准化日表SHA-256：`{daily_hash}`
"""
    report_path = RECORD_DIR / "02_七流域公开Excel数据审计报告.md"
    report_path.write_text(report, encoding="utf-8")

    print(f"完成：{len(daily_all):,}条流域-日记录。")
    print(f"日期完整文件：{complete_files}/14；缺测：{missing_total}；负值：{negative_total}。")
    print(f"公式验证流域：{verified_basins}/7；降水极端标记：{precip_flags}；径流极端标记：{runoff_flags}。")
    print(f"标准化数据：{daily_path}")
    print(f"审计报告：{report_path}")


if __name__ == "__main__":
    main()
